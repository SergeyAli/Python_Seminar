import openpyxl
# Define variable to load the wookbook
wookbook = openpyxl.load_workbook(r"C:\GB\python\20221013\Python_Seminar\Python_Seminar\Sem_7_D\sales.xlsx")
# Define variable to read the active sheet:
worksheet = wookbook.active
# Iterate the loop to read the cell values
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        print(col[i].value, end="\t\t")
    print('')